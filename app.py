"""Lead Extractor · Revolução AI"""
import os, io, csv, time, json
import streamlit as st
from dotenv import load_dotenv
load_dotenv()

def _s(k, d=""):
    try:
        v = st.secrets.get(k,"")
        if v: return str(v).strip()
    except: pass
    return os.getenv(k,d).strip()

# ── OAuth callback ────────────────────────────────────────────
_code  = st.query_params.get("code","")
_state = st.query_params.get("state","")
if _code and "sheets_creds" not in st.session_state:
    # 1) Credenciais vêm do parâmetro `state` (codificadas ao gerar a URL)
    #    — funciona independentemente de cookie ou sessão.
    cid, cs, ru = "", "", ""
    if _state:
        from modules.google_sheets import extrair_credenciais_state
        cid, cs, ru = extrair_credenciais_state(_state)

    # 2) Fallback: lê do env (state tem prioridade, não depende de sessão)
    if not (cid and cs):
        cid = cid or _s("GOOGLE_CLIENT_ID")
        cs  = cs  or _s("GOOGLE_CLIENT_SECRET")
        ru  = ru  or _s("APP_URL","http://localhost:8501")

    if cid and cs:
        try:
            from modules.google_sheets import trocar_codigo
            creds = trocar_codigo(cid, cs, ru, _code)
            st.session_state["sheets_creds"] = creds
            # Persiste no Supabase. Se sessão ainda não restaurada, salva depois em main().
            if "user" in st.session_state:
                from modules.database import salvar_configuracoes
                salvar_configuracoes({"google_sheets_creds": creds})
            else:
                st.session_state["_pending_sheets_save"] = True
        except Exception as e:
            st.session_state["_oauth_err"] = str(e)
    else:
        st.session_state["_oauth_err"] = "Credenciais OAuth não encontradas. Salve Client ID e Secret nas Configurações antes de conectar."
    st.query_params.clear()
    st.rerun()

st.set_page_config(page_title="Lead Extractor · Revolução AI", page_icon="⚡", layout="wide", initial_sidebar_state="expanded")

# ── CSS ───────────────────────────────────────────────────────
st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800&display=swap');

/* ── Reset & Base ─────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #08090E; }
::-webkit-scrollbar-thumb { background: #1C2235; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #00D97E40; }

/* ── Layout ───────────────────────────────────────────────── */
.stApp { background: #07080D !important; }
.block-container { padding: 2rem 2.5rem 4rem !important; max-width: 1280px !important; }
[data-testid="stAppViewContainer"] { background: #07080D; }

/* ── Sidebar ──────────────────────────────────────────────── */
[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #0A0C14 0%, #070810 100%) !important;
  border-right: 1px solid #141828 !important;
}
[data-testid="stSidebarContent"] { padding: 0 !important; }
[data-testid="stSidebar"] hr { margin: 0 16px; border-color: #141828 !important; }
[data-testid="stSidebar"] [data-testid="stButton"] > button {
  width: 100% !important; text-align: left !important;
  justify-content: flex-start !important; padding: 10px 16px !important;
  border-radius: 10px !important; font-size: 0.875rem !important;
  font-weight: 500 !important; transition: all 0.15s ease !important;
  margin: 1px 0 !important; border: none !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] > button[kind="secondary"] {
  background: transparent !important; color: #6A7490 !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] > button[kind="secondary"]:hover {
  background: #0F1420 !important; color: #C0C8DC !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] > button[kind="primary"] {
  background: linear-gradient(135deg, #0D2018, #0A1810) !important;
  color: #00D97E !important; border: 1px solid #00D97E25 !important;
  font-weight: 600 !important;
}

/* ── Inputs & Selects ─────────────────────────────────────── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {
  background: #0C0F1A !important; border: 1px solid #181E30 !important;
  border-radius: 10px !important; color: #E0E4F0 !important;
  padding: 10px 14px !important; font-size: 0.9rem !important;
  transition: border-color 0.2s, box-shadow 0.2s !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus {
  border-color: #00D97E55 !important; box-shadow: 0 0 0 3px #00D97E12 !important;
  outline: none !important;
}
[data-testid="stTextInput"] input::placeholder,
[data-testid="stNumberInput"] input::placeholder { color: #3A4255 !important; }
[data-testid="stSelectbox"] > div > div,
[data-testid="stMultiSelect"] > div > div {
  background: #0C0F1A !important; border: 1px solid #181E30 !important;
  border-radius: 10px !important; color: #E0E4F0 !important;
}

/* ── Forms ────────────────────────────────────────────────── */
[data-testid="stForm"] {
  background: linear-gradient(160deg, #0C0F1C 0%, #090B15 100%) !important;
  border: 1px solid #161C2E !important; border-radius: 16px !important;
  padding: 24px 26px !important;
  box-shadow: 0 8px 32px #00000040, inset 0 1px 0 #FFFFFF06 !important;
}

/* ── Buttons ──────────────────────────────────────────────── */
[data-testid="stButton"] > button,
[data-testid="stDownloadButton"] > button {
  border-radius: 8px !important; font-weight: 500 !important;
  font-size: 0.845rem !important; transition: all 0.18s ease !important;
  letter-spacing: 0.005em !important; padding: 7px 16px !important;
  min-height: unset !important; height: auto !important;
}
[data-testid="stButton"] > button[kind="primary"],
button[kind="primaryFormSubmit"] {
  background: #00C472 !important;
  color: #030E06 !important; border: none !important;
  font-weight: 600 !important;
  box-shadow: 0 2px 12px #00C47230 !important;
  padding: 9px 20px !important;
}
[data-testid="stButton"] > button[kind="primary"]:hover,
button[kind="primaryFormSubmit"]:hover {
  background: #00D97E !important;
  box-shadow: 0 4px 20px #00D97E45 !important;
  transform: translateY(-1px) !important;
}
[data-testid="stButton"] > button[kind="secondary"] {
  background: transparent !important; border: 1px solid #1C2438 !important;
  color: #6878A0 !important;
}
[data-testid="stButton"] > button[kind="secondary"]:hover {
  background: #0D1220 !important; border-color: #283050 !important;
  color: #C0C8DC !important;
}
[data-testid="stDownloadButton"] > button {
  background: transparent !important; border: 1px solid #1C2438 !important;
  color: #6878A0 !important;
}
[data-testid="stDownloadButton"] > button:hover {
  background: #0A1410 !important; border-color: #1A3C28 !important;
  color: #00C472 !important;
}

/* ── Tabs ─────────────────────────────────────────────────── */
[data-testid="stTabs"] [role="tablist"] {
  background: #0A0C16 !important; border-radius: 12px !important;
  padding: 5px !important; border: 1px solid #141828 !important; gap: 3px;
}
[data-testid="stTabs"] [role="tab"] {
  border-radius: 9px !important; color: #505870 !important;
  font-weight: 500 !important; font-size: 0.875rem !important;
  padding: 9px 22px !important; transition: all 0.2s !important;
  border: none !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
  background: linear-gradient(135deg, #0D2018, #091610) !important;
  color: #00D97E !important; font-weight: 600 !important;
  box-shadow: 0 2px 10px #00000030, inset 0 1px 0 #00D97E15 !important;
}
[data-testid="stTabs"] [role="tab"]:hover:not([aria-selected="true"]) {
  background: #0F1220 !important; color: #8090B0 !important;
}

/* ── Expanders ────────────────────────────────────────────── */
[data-testid="stExpander"] {
  background: #0A0D18 !important; border: 1px solid #141828 !important;
  border-radius: 14px !important; margin-bottom: 10px !important;
  overflow: hidden !important; transition: border-color 0.2s !important;
}
[data-testid="stExpander"]:hover { border-color: #1E2840 !important; }
[data-testid="stExpander"] summary {
  padding: 14px 20px !important; color: #C0C8DC !important;
  font-weight: 500 !important; font-size: 0.9rem !important;
}
[data-testid="stExpander"] summary:hover { background: #0D1020 !important; color: #E8EDF8 !important; }
[data-testid="stExpander"] > div > div { padding: 0 20px 18px !important; }

/* ── Alerts ───────────────────────────────────────────────── */
[data-testid="stAlert"] { border-radius: 12px !important; border-width: 1px !important; }
.stSuccess { background: #071510 !important; border-color: #00D97E30 !important; }
.stInfo { background: #050C1C !important; border-color: #2060C030 !important; }
.stWarning { background: #140F02 !important; border-color: #FFB80030 !important; }
.stError { background: #140404 !important; border-color: #FF475730 !important; }

/* ── Progress ─────────────────────────────────────────────── */
[data-testid="stProgress"] { margin: 12px 0 !important; }
[data-testid="stProgress"] > div { background: #10141E !important; border-radius: 6px !important; height: 6px !important; }
[data-testid="stProgress"] > div > div {
  background: linear-gradient(90deg, #00D97E, #00F090) !important;
  border-radius: 6px !important; box-shadow: 0 0 10px #00D97E60 !important;
}

/* ── DataFrames ───────────────────────────────────────────── */
.stDataFrame { border-radius: 12px !important; overflow: hidden !important; }
.stDataFrame [data-testid="stDataFrameResizable"] { border: 1px solid #141828 !important; border-radius: 12px !important; }

/* ── Toggle ───────────────────────────────────────────────── */
[data-testid="stToggle"] { background: transparent !important; padding: 8px 0 !important; }

/* ── Captions & Labels ────────────────────────────────────── */
[data-testid="stCaptionContainer"] { color: #4A5570 !important; }
[data-testid="stWidgetLabel"] { color: #6A7490 !important; font-size: 0.82rem !important; font-weight: 500 !important; }

/* ══════════════════════════════════════════════════════════
   CUSTOM COMPONENTS
══════════════════════════════════════════════════════════ */

/* ── Sidebar brand ────────────────────────────────────────── */
.brand-header {
  padding: 22px 20px 18px; margin-bottom: 4px;
  background: linear-gradient(160deg, #0D1820 0%, #080A12 100%);
  border-bottom: 1px solid #141828;
}
.brand-dot {
  width: 36px; height: 36px;
  background: linear-gradient(135deg, #00D97E, #00A85E);
  border-radius: 10px; display: inline-flex; align-items: center;
  justify-content: center; font-weight: 800; color: #040A06;
  font-size: 16px; box-shadow: 0 4px 16px #00D97E45;
  vertical-align: middle; margin-right: 10px; flex-shrink: 0;
}
.brand-title {
  font-size: 0.92rem; font-weight: 700; color: #E8EDF8;
  line-height: 1.2; letter-spacing: -0.01em;
}
.brand-sub {
  font-size: 0.62rem; color: #00D97E; font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.14em; margin-top: 2px;
}

/* ── User card in sidebar ─────────────────────────────────── */
.user-card {
  margin: 8px 12px 4px;
  background: #0C1020; border: 1px solid #141828;
  border-radius: 12px; padding: 10px 14px;
}
.user-email { font-size: 0.78rem; color: #7080A0; font-weight: 500; word-break: break-all; }
.user-role-badge {
  display: inline-block; margin-top: 5px; padding: 2px 8px;
  border-radius: 6px; font-size: 0.65rem; font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.1em;
}
.role-admin { background: #0A1E14; color: #00D97E; border: 1px solid #00D97E20; }
.role-user  { background: #0A0E1C; color: #6080C0; border: 1px solid #2040A030; }

/* ── Page header ──────────────────────────────────────────── */
.page-header {
  margin-bottom: 2rem; padding-bottom: 1.5rem;
  border-bottom: 1px solid #0E1220;
  display: flex; align-items: flex-start; gap: 14px;
}
.page-header-icon {
  width: 42px; height: 42px; border-radius: 11px; flex-shrink: 0;
  display: flex; align-items: center; justify-content: center;
  font-size: 20px; background: #0C1820; border: 1px solid #1A2A40;
}
.page-title {
  font-size: 1.55rem; font-weight: 800; letter-spacing: -0.025em;
  line-height: 1.15;
  background: linear-gradient(135deg, #E8EDF8 20%, #8090B0 100%);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
  background-clip: text; margin-bottom: 0.2rem;
}
.page-sub { font-size: 0.855rem; color: #485068; font-weight: 400; }

/* ── Stat cards ───────────────────────────────────────────── */
.stats-row { display: grid; grid-template-columns: repeat(4,1fr); gap: 14px; margin: 1.25rem 0; }
.stat-card {
  background: linear-gradient(160deg, #0C1020 0%, #090C18 100%);
  border: 1px solid #141828; border-radius: 16px; padding: 18px 20px;
  position: relative; overflow: hidden;
  transition: border-color 0.2s, transform 0.2s, box-shadow 0.2s;
}
.stat-card:hover { border-color: #00D97E25; transform: translateY(-2px); box-shadow: 0 8px 32px #00000050; }
.stat-card::before {
  content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px;
  background: linear-gradient(90deg, #00D97E, transparent);
}
.stat-num { font-size: 2.1rem; font-weight: 800; color: #00D97E; line-height: 1; letter-spacing: -0.03em; }
.stat-lbl { font-size: 0.7rem; color: #485068; margin-top: 6px; text-transform: uppercase; letter-spacing: 0.1em; font-weight: 600; }

/* ── Section label ────────────────────────────────────────── */
.sec {
  font-size: 0.68rem; font-weight: 700; color: #00D97E;
  text-transform: uppercase; letter-spacing: 0.13em;
  margin: 1.1rem 0 0.45rem;
  display: flex; align-items: center; gap: 8px;
}
.sec::after {
  content: ''; flex: 1; height: 1px;
  background: linear-gradient(90deg, #00D97E20, transparent);
}

/* ── Info box ─────────────────────────────────────────────── */
.info-box {
  background: linear-gradient(135deg, #090F14, #060C10);
  border: 1px solid #14202A; border-left: 3px solid #00D97E;
  border-radius: 10px; padding: 12px 16px;
  font-size: 0.84rem; color: #607888; line-height: 1.65; margin-bottom: 1.25rem;
}
.info-box strong { color: #90B0A0; }

/* ── Divider ──────────────────────────────────────────────── */
.hr { border: none; border-top: 1px solid #0E1220; margin: 1rem 0; }

/* ── Badges ───────────────────────────────────────────────── */
.badge { display: inline-flex; align-items: center; gap: 4px; padding: 4px 10px; border-radius: 20px; font-size: 0.75rem; font-weight: 600; }
.b-ok  { background: #061510; color: #00D97E; border: 1px solid #00D97E25; }
.b-warn{ background: #140E00; color: #FFB800; border: 1px solid #FFB80025; }
.b-err { background: #140404; color: #FF4757; border: 1px solid #FF475725; }

/* ── Login form override ──────────────────────────────────── */
/* The login form is a single unified card: logo + title + fields */
[data-testid="stForm"] { padding: 28px 28px 24px !important; }
.login-head { text-align: center; padding: 2px 0 22px; }
.login-title {
  font-size: 1.3rem; font-weight: 700; color: #E0E6F5;
  letter-spacing: -0.02em; margin-top: 14px;
}
.login-sub {
  font-size: 0.66rem; color: #4A8C6A; font-weight: 600;
  letter-spacing: 0.13em; text-transform: uppercase; margin-top: 3px;
}

</style>""", unsafe_allow_html=True)

from modules.google_sheets import COLUNAS_EXPORT
from modules.auth import _COOKIE_NAME  # nome do cookie de sessão

def _logo_html(size: int = 40) -> str:
    """Retorna HTML com a logo.
    Usa URL estática (app/static/logo.png) — o browser faz cache da imagem e
    ela NÃO é reenviada via WebSocket a cada render, ao contrário do base64."""
    from pathlib import Path
    p = Path("static/logo.png")
    r = size // 5
    if p.exists():
        return (f'<img src="app/static/logo.png" '
                f'style="width:{size}px;height:{size}px;border-radius:{r}px;'
                f'object-fit:cover;display:inline-block" />')
    r2 = size // 4
    fs = int(size * 0.38)
    return (f'<div style="width:{size}px;height:{size}px;background:#00C472;'
            f'border-radius:{r2}px;display:inline-flex;align-items:center;'
            f'justify-content:center;font-size:{fs}px;font-weight:800;color:#030E06">R</div>')
_EXTRA=[("telefone_internacional","Telefone Intl."),("status_funcionamento","Status"),
        ("porte","Porte"),("data_abertura","Data Abertura")]
ALL_COLS = COLUNAS_EXPORT + [c for c in _EXTRA if c not in COLUNAS_EXPORT]

def _csv(rows):
    b=io.StringIO(); w=csv.writer(b)
    w.writerow([l for _,l in ALL_COLS])
    for r in rows: w.writerow([r.get(c,"") for c,_ in ALL_COLS])
    return b.getvalue().encode("utf-8-sig")

def _xlsx(rows):
    import openpyxl
    from openpyxl.styles import Font,PatternFill,Alignment
    from openpyxl.utils import get_column_letter
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Prospecção"
    hf=Font(bold=True,color="0A0A0F"); hfill=PatternFill("solid",fgColor="00D97E")
    for ci,(col,lbl) in enumerate(ALL_COLS,1):
        c=ws.cell(row=1,column=ci,value=lbl); c.font=hf; c.fill=hfill
        c.alignment=Alignment(horizontal="center")
    ws.row_dimensions[1].height=22
    fa=PatternFill("solid",fgColor="141418"); fb=PatternFill("solid",fgColor="0D0D12")
    for ri,r in enumerate(rows,2):
        for ci,(col,_) in enumerate(ALL_COLS,1):
            cell=ws.cell(row=ri,column=ci,value=r.get(col,""))
            cell.fill=fa if ri%2==0 else fb
            if col in("site","maps_url") and r.get(col):
                cell.hyperlink=r[col]; cell.font=Font(color="00D97E",underline="single")
    for ci in range(1,len(ALL_COLS)+1):
        mx=max(len(str(ws.cell(row=rr,column=ci).value or"")) for rr in range(1,len(rows)+2))
        ws.column_dimensions[get_column_letter(ci)].width=min(mx+2,50)
    ws.freeze_panes="A2"
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def _stats(rows):
    tot  = len(rows)
    tel  = sum(1 for r in rows if r.get("telefone") or r.get("telefone2"))
    site = sum(1 for r in rows if r.get("site"))
    em   = sum(1 for r in rows if r.get("email"))
    cards = "".join(
        f'<div class="stat-card"><div class="stat-num">{n}</div><div class="stat-lbl">{l}</div></div>'
        for n, l in [(tot,"Total leads"),(tel,"Com telefone"),(site,"Com site"),(em,"Com e-mail")]
    )
    st.markdown(f'<div class="stats-row">{cards}</div>', unsafe_allow_html=True)

def _tabela(rows):
    import pandas as pd
    vis=["nome","telefone","email","municipio","uf","endereco","site","avaliacao","cnpj","nicho_busca","subnicho_busca","fonte"]
    lm={c:l for c,l in ALL_COLS}; df=pd.DataFrame(rows)
    cols=[c for c in vis if c in df.columns]
    st.dataframe(df[cols].rename(columns=lm).fillna("").astype(str).replace("nan",""), use_container_width=True, height=380)

def _dl_buttons(rows, prefix, sheets_auth):
    ts=int(time.time()); c1,c2,c3=st.columns(3)
    with c1: st.download_button("⬇️ Excel",_xlsx(rows),f"{prefix}_{ts}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    with c2: st.download_button("⬇️ CSV",_csv(rows),f"{prefix}_{ts}.csv","text/csv",use_container_width=True)
    with c3:
        if sheets_auth:
            if st.button("📊 Google Sheets",use_container_width=True): _export_sheets(rows)
        else:
            st.button("📊 Google Sheets",use_container_width=True,disabled=True,help="Conecte sua conta Google em Configurações.")

def _export_sheets(rows):
    from modules.google_sheets import exportar
    creds = st.session_state.get("sheets_creds")
    sid   = st.session_state.get("sheets_selected_id", "")
    aba   = st.session_state.get("sheets_aba", "Prospecção")
    modo  = st.session_state.get("sheets_modo", "substituir")
    if not creds:
        st.warning("Conecte sua conta Google em ⚙️ Configurações.", icon="🔗")
        return
    if not sid:
        st.warning("Selecione uma planilha destino em ⚙️ Configurações.", icon="📄")
        return
    with st.spinner("Exportando..."):
        ok, msg = exportar(rows, creds, sid, aba, modo)
    (st.success if ok else st.error)(msg)

# ══════════════════════════════════════════════════════════════
# PÁGINAS
# ══════════════════════════════════════════════════════════════

def pagina_login():
    # Centralização vertical: espaçador + colunas
    st.markdown("<div style='height:12vh'></div>", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("login"):
            st.markdown(
                f'<div class="login-head">'
                f'{_logo_html(52)}'
                f'<div class="login-title">Lead Extractor</div>'
                f'<div class="login-sub">Revolução AI</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
            email = st.text_input("E-mail", placeholder="seu@email.com")
            senha = st.text_input("Senha", type="password", placeholder="••••••••")
            btn   = st.form_submit_button("Entrar", use_container_width=True, type="primary")
        if btn:
            if not email or not senha:
                st.error("Preencha e-mail e senha.")
            else:
                from modules.auth import login, supabase_configurado
                if not supabase_configurado():
                    st.error("Supabase não configurado. Verifique SUPABASE_URL e SUPABASE_ANON_KEY nos Secrets.")
                else:
                    with st.spinner(""):
                        ok, msg = login(email, senha)
                    if ok:
                        st.session_state["page"] = "busca"
                        st.rerun()
                    else:
                        st.error(msg)


def pagina_busca():
    from modules.nichos import NICHOS, ESTADOS, NOMES_NICHOS, SIGLAS_ESTADOS

    st.markdown(
        '<div class="page-header">'
        '<div class="page-header-icon">🔍</div>'
        '<div><div class="page-title">Nova Busca</div>'
        '<div class="page-sub">Busque leads por nicho e localidade usando Google Maps ou Receita Federal</div></div>'
        '</div>',
        unsafe_allow_html=True,
    )

    # Chave do Google Maps vem EXCLUSIVAMENTE das configurações do usuário (Supabase)
    from modules.database import carregar_configuracoes
    _cfg_busca = carregar_configuracoes()
    gmaps_key = _cfg_busca.get("google_maps_api_key", "") or st.session_state.get("user_gmaps_key", "")
    gmaps_ok  = bool(gmaps_key)

    aba_maps, aba_rf = st.tabs(["🗺️  Google Maps  ·  com telefone", "📋  Receita Federal  ·  dados oficiais"])

    with aba_maps:
        if not gmaps_ok:
            st.warning("Chave do Google Maps não configurada. Acesse **Configurações → Google Maps API** para adicionar.", icon="⚠️")
        st.markdown('<div class="info-box">Melhor fonte para <strong>telefones</strong>. Até ~500 resultados com múltiplas buscas automáticas.</div>', unsafe_allow_html=True)

        col_n, col_s = st.columns(2)
        with col_n:
            st.markdown('<div class="sec">Nicho *</div>', unsafe_allow_html=True)
            nicho_sel = st.selectbox("Nicho", NOMES_NICHOS, label_visibility="collapsed", key="m_nicho")
        nicho_data = NICHOS[nicho_sel]; is_custom = nicho_sel == "Outro / Personalizado"
        with col_s:
            st.markdown('<div class="sec">Subnicho / Especialidade</div>', unsafe_allow_html=True)
            if is_custom:
                query_custom = st.text_input("Termo", placeholder='"pet shop", "clínica veterinária"', label_visibility="collapsed", key="m_qcustom")
                subnicho_sel = ""
            else:
                sub_opts = ["Todos (sem filtro)"] + nicho_data["subnichos"] + ["✏️ Personalizado..."]
                subnicho_sel = st.selectbox("Subnicho", sub_opts, label_visibility="collapsed", key="m_sub")
        sub_custom = ""
        if not is_custom and subnicho_sel == "✏️ Personalizado...":
            sub_custom = st.text_input("Especialidade personalizada", key="m_subcustom")

        st.markdown('<hr class="hr">', unsafe_allow_html=True)
        with st.form("form_maps"):
            st.markdown('<div class="sec">Localidade — cidade e/ou estado</div>', unsafe_allow_html=True)
            cc, ce, cl = st.columns([3,1,2])
            with cc: cidade = st.text_input("Cidade", placeholder="Ex: São Paulo", label_visibility="collapsed")
            with ce:
                eopts = ["—"]+SIGLAS_ESTADOS; edef = eopts.index("SP") if "SP" in eopts else 0
                est_raw = st.selectbox("Estado", eopts, index=edef, label_visibility="collapsed")
                estado = "" if est_raw == "—" else est_raw
            with cl:
                lim = st.slider("Resultados", 20, 500, 60, 20, label_visibility="collapsed")
                st.caption(f"Máx. **{lim}** resultados")
            apenas_novos_maps = st.toggle(
                "🔄 Apenas leads novos (remover repetidos de buscas anteriores)",
                value=True,
                help="Quando ativado, leads com mesmo telefone ou CNPJ de pesquisas anteriores são removidos dos resultados.",
            )
            buscar_btn = st.form_submit_button("🔍 Buscar no Google Maps", disabled=not gmaps_ok, use_container_width=True, type="primary")

        if buscar_btn:
            cv, ev = cidade.strip(), estado.strip()
            if not cv and not ev: st.error("Informe ao menos a cidade ou o estado.")
            elif is_custom and not query_custom.strip(): st.error("Informe o termo personalizado.")
            else:
                from modules.google_maps import buscar as maps_buscar
                qbase = query_custom.strip() if is_custom else nicho_data["query"]
                nicho_lbl = qbase if is_custom else nicho_sel
                sub_final = "" if (is_custom or subnicho_sel=="Todos (sem filtro)") else (sub_custom.strip() if subnicho_sel=="✏️ Personalizado..." else subnicho_sel)
                localidade = f"{cv}, {ESTADOS.get(ev,ev)}" if cv and ev else cv or ESTADOS.get(ev,ev)
                slug = f"{nicho_lbl[:15]}_{localidade[:15]}".lower().replace(" ","_").replace(",","")
                # Carrega identificadores já salvos antes de iniciar a busca
                excl_tels_maps = set()
                if apenas_novos_maps:
                    from modules.database import buscar_identificadores_existentes
                    excl_tels_maps, _ = buscar_identificadores_existentes()
                prog = st.progress(0, text="Iniciando...")
                def _cb(a, t, m):
                    v = min(a / t, 1.0) if t and t > 0 else 0
                    prog.progress(v, text=str(m)[:120])
                try:
                    res = maps_buscar(query_base=qbase, localidade=localidade, limite=lim,
                                      api_key=gmaps_key, nicho=nicho_lbl, subnicho=sub_final,
                                      cidade=cv, estado=ev, progress_callback=_cb,
                                      exclude_phones=excl_tels_maps if apenas_novos_maps else None)
                    prog.progress(1.0, text=f"Concluído! {len(res)} resultados.")
                    prog.empty()
                    st.session_state["maps_res"] = res
                    st.session_state["maps_prefix"] = slug
                except ValueError as e:
                    prog.empty(); st.error(str(e)); st.session_state["maps_res"] = []
                except Exception as e:
                    prog.empty(); st.error(f"Erro: {e}"); st.session_state["maps_res"] = []
                else:
                    # Salva no banco em background — não bloqueia exibição de resultados
                    try:
                        from modules.database import salvar_pesquisa, salvar_leads
                        sid = salvar_pesquisa(nicho_lbl, sub_final, cv, ev, localidade, "maps", len(res))
                        if sid: salvar_leads(sid, res)
                    except Exception:
                        pass  # falha no banco não apaga os resultados

        if st.session_state.get("maps_res"):
            res = st.session_state["maps_res"]
            st.success(f"✅ **{len(res)}** resultados")
            _stats(res); _dl_buttons(res, st.session_state.get("maps_prefix","prospecao"), "sheets_creds" in st.session_state)
            st.markdown("#### Prévia"); _tabela(res)

    with aba_rf:
        st.markdown('<div class="info-box">Dados oficiais da <strong>Receita Federal</strong>. Ideal para volumes maiores com CNPJ e e-mail. <strong>1º uso:</strong> baixa arquivos ~350 MB (fica em cache 30 dias).</div>', unsafe_allow_html=True)
        with st.form("form_rf"):
            c1,c2=st.columns(2)
            with c1: mun_rf=st.text_input("Município (opcional)", placeholder="Ex: São Paulo", label_visibility="collapsed")
            with c2: uf_rf=st.selectbox("Estado *", SIGLAS_ESTADOS, index=SIGLAS_ESTADOS.index("SP"), label_visibility="collapsed")
            lim_rf=st.slider("Máx. resultados",50,2000,300,50)
            apenas_novos_rf = st.toggle(
                "🔄 Apenas leads novos (remover repetidos de buscas anteriores)",
                value=True,
                help="Quando ativado, empresas com mesmo CNPJ ou telefone de pesquisas anteriores são removidas.",
            )
            btn_rf=st.form_submit_button("🔍 Buscar na Receita Federal", use_container_width=True, type="primary")
        if btn_rf:
            from modules.receita_federal import buscar_por_cnae_rf
            local = mun_rf.strip() or uf_rf
            # Carrega identificadores já salvos antes de iniciar a busca
            excl_tels_rf, excl_cnpjs_rf = set(), set()
            if apenas_novos_rf:
                from modules.database import buscar_identificadores_existentes
                excl_tels_rf, excl_cnpjs_rf = buscar_identificadores_existentes()
            bar = st.progress(0, text="Iniciando...")
            def _cbrf(a, t, m):
                v = min(a / t, 1.0) if t and t > 0 else 0
                bar.progress(v, text=str(m)[:120])
            try:
                res_rf = buscar_por_cnae_rf(
                    uf=uf_rf, municipio=mun_rf.strip(), limite=lim_rf,
                    callback_progresso=_cbrf,
                    exclude_phones=excl_tels_rf if apenas_novos_rf else None,
                    exclude_cnpjs=excl_cnpjs_rf if apenas_novos_rf else None,
                )
                bar.progress(1.0, text=f"Concluído! {len(res_rf)} resultados.")
                bar.empty()
                st.session_state["rf_res"] = res_rf
                st.session_state["rf_prefix"] = f"rf_{local.lower().replace(' ','_')}"
            except Exception as e:
                bar.empty(); st.error(f"Erro: {e}"); st.session_state["rf_res"] = []
            else:
                try:
                    from modules.database import salvar_pesquisa, salvar_leads
                    sid = salvar_pesquisa("Advocacia (RF)", "", mun_rf.strip(), uf_rf, local, "receita_federal", len(res_rf))
                    if sid: salvar_leads(sid, res_rf)
                except Exception:
                    pass
        if st.session_state.get("rf_res"):
            res=st.session_state["rf_res"]
            st.success(f"✅ **{len(res)}** resultados")
            _stats(res); _dl_buttons(res,st.session_state.get("rf_prefix","prospecao_rf"),"sheets_creds" in st.session_state)
            st.markdown("#### Prévia"); _tabela(res)


def pagina_historico():
    from modules.database import listar_pesquisas, buscar_leads_da_pesquisa, deletar_pesquisa

    st.markdown(
        '<div class="page-header">'
        '<div class="page-header-icon">📁</div>'
        '<div><div class="page-title">Histórico</div>'
        '<div class="page-sub">Todas as suas extrações anteriores com leads salvos</div></div>'
        '</div>',
        unsafe_allow_html=True,
    )

    pesquisas = listar_pesquisas()
    if not pesquisas:
        st.info("Nenhuma pesquisa salva ainda. Faça sua primeira busca!", icon="💡")
        return

    for p in pesquisas:
        dt = p.get("created_at","")[:16].replace("T"," ") if p.get("created_at") else "—"
        fonte_icon = "🗺️" if p.get("fonte") == "maps" else "📋"
        nicho = p.get("nicho","—"); sub = p.get("subnicho",""); loc = p.get("localidade","—")
        total = p.get("total_results", 0)
        titulo = f"{fonte_icon} **{nicho}**" + (f" · {sub}" if sub else "") + f" — {loc}"

        with st.expander(f"{titulo}  ·  {total} leads  ·  {dt}"):
            col_a, col_b = st.columns([6,1])
            with col_b:
                if st.button("🗑️ Apagar", key=f"del_{p['id']}"):
                    ok, msg = deletar_pesquisa(p["id"])
                    (st.success if ok else st.error)(msg)
                    if ok: time.sleep(0.5); st.rerun()

            leads = buscar_leads_da_pesquisa(p["id"])
            if not leads:
                st.caption("Nenhum lead salvo para esta pesquisa.")
                continue

            ts = int(time.time())
            slug = f"{nicho[:12]}_{loc[:12]}".lower().replace(" ","_").replace(",","")
            c1,c2 = st.columns(2)
            with c1: st.download_button("⬇️ Excel",_xlsx(leads),f"{slug}_{ts}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,key=f"xl_{p['id']}")
            with c2: st.download_button("⬇️ CSV",_csv(leads),f"{slug}_{ts}.csv","text/csv",use_container_width=True,key=f"csv_{p['id']}")

            import pandas as pd
            vis=["nome","telefone","email","municipio","uf","site","avaliacao","cnpj"]
            lm={c:l for c,l in ALL_COLS}; df=pd.DataFrame(leads)
            st.dataframe(df[[c for c in vis if c in df.columns]].rename(columns=lm).fillna("").astype(str).replace("nan",""), use_container_width=True, height=280)


# ── Configurações ──────────────────────────────────────────────────────────────

def pagina_configuracoes():
    from modules.database import carregar_configuracoes, salvar_configuracoes
    from modules.google_sheets import gerar_url_auth

    st.markdown(
        '<div class="page-header">'
        '<div class="page-header-icon">⚙️</div>'
        '<div><div class="page-title">Configurações</div>'
        '<div class="page-sub">Gerencie suas credenciais e integrações</div></div>'
        '</div>',
        unsafe_allow_html=True,
    )

    cfg = carregar_configuracoes()

    # ── Google Maps ─────────────────────────────────────────────────────────────
    with st.expander("🗺️ Google Maps API", expanded=True):
        st.markdown("Insira sua chave de API do Google Maps (Places API).")
        gmk = st.text_input(
            "Chave de API",
            value=cfg.get("google_maps_api_key",""),
            type="password",
            placeholder="AIzaSy...",
            key="cfg_gmaps",
        )
        if st.button("💾 Salvar chave Maps", key="save_gmaps"):
            ok, msg = salvar_configuracoes({"google_maps_api_key": gmk})
            (st.success if ok else st.error)(msg)
            if ok:
                st.session_state["user_gmaps_key"] = gmk

    # ── Google Sheets OAuth ─────────────────────────────────────────────────────
    with st.expander("📊 Google Sheets (OAuth)", expanded=True):
        st.markdown("Conecte sua conta Google para exportar resultados diretamente para planilhas.")

        # Credenciais OAuth do app
        cid = cfg.get("google_client_id","") or _s("GOOGLE_CLIENT_ID")
        cs  = cfg.get("google_client_secret","") or _s("GOOGLE_CLIENT_SECRET")
        ru  = cfg.get("app_url","") or _s("APP_URL","http://localhost:8501")

        col1, col2 = st.columns(2)
        with col1:
            new_cid = st.text_input("Google Client ID",  value=cid, placeholder="...apps.googleusercontent.com", key="cfg_cid")
        with col2:
            new_cs  = st.text_input("Google Client Secret", value=cs, type="password", placeholder="GOCSPX-...", key="cfg_cs")

        new_ru = st.text_input("URL do app (para redirect OAuth)", value=ru, placeholder="https://seu-app.streamlit.app", key="cfg_ru")

        if st.button("💾 Salvar credenciais Google", key="save_google_creds"):
            ok, msg = salvar_configuracoes({
                "google_client_id": new_cid,
                "google_client_secret": new_cs,
                "app_url": new_ru,
            })
            (st.success if ok else st.error)(msg)

        st.markdown("---")

        if "sheets_creds" in st.session_state:
            st.success("✅ Conta Google conectada!", icon="✅")

            # ── Seleção de planilha ─────────────────────────────────────────
            from modules.google_sheets import listar_planilhas, listar_abas
            if "sheets_lista" not in st.session_state:
                with st.spinner("Buscando planilhas no Google Drive..."):
                    try:
                        st.session_state["sheets_lista"] = listar_planilhas(st.session_state["sheets_creds"])
                    except Exception as e:
                        st.session_state["sheets_lista"] = []
                        st.warning(f"Erro ao listar planilhas: {e}")

            planilhas = st.session_state.get("sheets_lista", [])
            if planilhas:
                nomes = [p["name"] for p in planilhas]
                ids   = [p["id"]   for p in planilhas]
                sel_name = st.session_state.get("sheets_selected_name", "")
                idx = nomes.index(sel_name) if sel_name in nomes else 0
                escolha = st.selectbox("📄 Planilha destino", nomes, index=idx, key="cfg_sheet_sel")
                chosen_id = ids[nomes.index(escolha)]

                if chosen_id != st.session_state.get("sheets_selected_id"):
                    st.session_state["sheets_selected_id"] = chosen_id
                    st.session_state["sheets_selected_name"] = escolha
                    st.session_state.pop("sheets_abas", None)

                if "sheets_abas" not in st.session_state:
                    with st.spinner("Carregando abas..."):
                        try:
                            st.session_state["sheets_abas"] = listar_abas(st.session_state["sheets_creds"], chosen_id)
                        except Exception:
                            st.session_state["sheets_abas"] = ["Prospecção"]

                abas = st.session_state.get("sheets_abas", ["Prospecção"])
                aba_atual = st.session_state.get("sheets_aba", "Prospecção")
                aba_idx = abas.index(aba_atual) if aba_atual in abas else 0

                col_aba, col_modo = st.columns(2)
                with col_aba:
                    aba_sel = st.selectbox("📑 Aba destino", abas, index=aba_idx, key="cfg_aba_sel")
                    st.session_state["sheets_aba"] = aba_sel
                with col_modo:
                    modo_opts = ["substituir", "acrescentar"]
                    modo_idx = modo_opts.index(st.session_state.get("sheets_modo", "substituir"))
                    modo_sel = st.selectbox("📝 Modo de exportação", modo_opts, index=modo_idx, key="cfg_modo_sel")
                    st.session_state["sheets_modo"] = modo_sel

                st.info(f"Exportação irá para **{escolha}** → aba **{aba_sel}** ({modo_sel})", icon="📊")
            else:
                st.warning("Nenhuma planilha encontrada. Crie uma planilha no Google Drive e recarregue.", icon="⚠️")

            st.markdown("")
            if st.button("🔓 Desconectar conta Google", key="disc_google"):
                st.session_state.pop("sheets_creds", None)
                st.session_state.pop("sheets_lista", None)
                st.session_state.pop("sheets_selected_id", None)
                st.session_state.pop("sheets_selected_name", None)
                st.session_state.pop("sheets_abas", None)
                st.session_state.pop("sheets_aba", None)
                st.session_state.pop("sheets_modo", None)
                salvar_configuracoes({"google_sheets_creds": None})
                st.rerun()
        else:
            if new_cid and new_cs:
                url = gerar_url_auth(new_cid, new_cs, new_ru)
                st.link_button("🔗 Conectar conta Google", url, use_container_width=True)
            else:
                st.info("Preencha o Client ID e Secret acima para conectar sua conta Google.", icon="ℹ️")

    # ── Alterar senha ────────────────────────────────────────────────────────────
    with st.expander("🔑 Alterar senha", expanded=False):
        np1 = st.text_input("Nova senha", type="password", key="cfg_np1")
        np2 = st.text_input("Confirmar nova senha", type="password", key="cfg_np2")
        if st.button("🔐 Alterar senha", key="btn_alterar_senha"):
            if not np1:
                st.warning("Digite a nova senha.")
            elif np1 != np2:
                st.error("As senhas não coincidem.")
            elif len(np1) < 6:
                st.error("A senha deve ter pelo menos 6 caracteres.")
            else:
                from modules.auth import redefinir_senha
                uid = st.session_state.get("user",{}).get("id","")
                ok, msg = redefinir_senha(uid, np1)
                (st.success if ok else st.error)(msg)


# ── Admin ──────────────────────────────────────────────────────────────────────

def pagina_admin():
    from modules.auth import (
        listar_usuarios, criar_usuario, deletar_usuario,
        alterar_role, redefinir_senha,
    )

    st.markdown(
        '<div class="page-header">'
        '<div class="page-header-icon">👑</div>'
        '<div><div class="page-title">Painel Admin</div>'
        '<div class="page-sub">Gerencie os usuários da plataforma</div></div>'
        '</div>',
        unsafe_allow_html=True,
    )

    # ── Criar usuário ────────────────────────────────────────────────────────────
    with st.expander("➕ Criar novo usuário", expanded=False):
        c1, c2, c3 = st.columns([3,2,1])
        with c1: new_email = st.text_input("E-mail", key="adm_email", placeholder="usuario@empresa.com")
        with c2: new_senha = st.text_input("Senha inicial", type="password", key="adm_senha", placeholder="Mín. 6 caracteres")
        with c3: new_role  = st.selectbox("Papel", ["user","admin"], key="adm_role")
        if st.button("✅ Criar usuário", key="btn_criar_user"):
            if not new_email or not new_senha:
                st.warning("Preencha e-mail e senha.")
            elif len(new_senha) < 6:
                st.error("Senha deve ter pelo menos 6 caracteres.")
            else:
                ok, msg = criar_usuario(new_email.strip(), new_senha, new_role)
                (st.success if ok else st.error)(msg)
                if ok: time.sleep(0.3); st.rerun()

    st.markdown("---")

    # ── Lista de usuários ────────────────────────────────────────────────────────
    ok, usuarios, err = listar_usuarios()
    if not ok:
        st.error(f"Não foi possível carregar usuários: {err}")
        return
    if not usuarios:
        st.info("Nenhum usuário cadastrado.")
        return

    st.markdown(f"**{len(usuarios)} usuário(s) cadastrado(s)**")

    for u in usuarios:
        uid       = u.get("id","")
        email     = u.get("email","—")
        role      = u.get("role","user")
        created   = (u.get("created_at","") or "")[:10]
        searches  = u.get("total_searches", 0) or 0
        leads_tot = u.get("total_leads", 0) or 0
        last_s    = (u.get("last_search_at","") or "")[:10] or "nunca"
        me        = st.session_state.get("user",{}).get("id","") == uid

        badge = "🟢 admin" if role == "admin" else "⚪ user"
        label = f"{badge}  **{email}**" + ("  *(você)*" if me else "")

        with st.expander(label):
            st.caption(f"ID: `{uid}`  ·  Criado em {created}  ·  {searches} pesquisas  ·  {leads_tot} leads  ·  Última busca: {last_s}")

            col_r, col_p, col_d = st.columns(3)

            with col_r:
                novo_role = st.selectbox(
                    "Papel", ["user","admin"],
                    index=0 if role == "user" else 1,
                    key=f"role_{uid}",
                )
                if st.button("🔄 Alterar papel", key=f"btn_role_{uid}", disabled=me):
                    ok2, msg2 = alterar_role(uid, novo_role)
                    (st.success if ok2 else st.error)(msg2)
                    if ok2: time.sleep(0.3); st.rerun()

            with col_p:
                nova_senha = st.text_input("Nova senha", type="password", key=f"pw_{uid}", placeholder="Mín. 6 caracteres")
                if st.button("🔑 Redefinir senha", key=f"btn_pw_{uid}"):
                    if not nova_senha or len(nova_senha) < 6:
                        st.warning("Mínimo 6 caracteres.")
                    else:
                        ok3, msg3 = redefinir_senha(uid, nova_senha)
                        (st.success if ok3 else st.error)(msg3)

            with col_d:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("🗑️ Remover usuário", key=f"del_u_{uid}", disabled=me, type="secondary"):
                    ok4, msg4 = deletar_usuario(uid)
                    (st.success if ok4 else st.error)(msg4)
                    if ok4: time.sleep(0.3); st.rerun()


# ── Sidebar & roteamento principal ────────────────────────────────────────────

def _sidebar():
    from modules.auth import logout, eh_admin

    with st.sidebar:
        # ── Brand ──────────────────────────────────────────────
        user = st.session_state.get("user", {})
        role = user.get("role", "user")
        logo = _logo_html(30)
        st.markdown(
            f'<div class="brand-header">'
            f'<div style="display:flex;align-items:center;gap:10px">'
            f'{logo}'
            f'<div><div class="brand-title">Lead Extractor</div>'
            f'<div class="brand-sub">Revolução AI</div></div>'
            f'</div></div>',
            unsafe_allow_html=True,
        )

        # ── User card ───────────────────────────────────────────
        role_cls   = "role-admin" if role == "admin" else "role-user"
        role_label = "Admin" if role == "admin" else "Usuário"
        st.markdown(
            f'<div class="user-card">'
            f'<div class="user-email">{user.get("email","")}</div>'
            f'<span class="user-role-badge {role_cls}">{role_label}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

        # ── Nav — sem emojis, texto limpo ────────────────────────
        page = st.session_state.get("page", "busca")
        nav_items = [
            ("busca",         "Busca"),
            ("historico",     "Histórico"),
            ("configuracoes", "Configurações"),
        ]
        if eh_admin():
            nav_items.append(("admin", "Admin"))

        st.markdown('<div style="padding:0 8px">', unsafe_allow_html=True)
        for key, label in nav_items:
            if st.button(label, use_container_width=True, key=f"nav_{key}",
                         type="primary" if page == key else "secondary"):
                st.session_state["page"] = key
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        # ── Logout ──────────────────────────────────────────────
        st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)
        st.markdown('<hr style="border-color:#141828;margin:0 8px 12px">', unsafe_allow_html=True)
        st.markdown('<div style="padding:0 8px">', unsafe_allow_html=True)
        if st.button("Sair", use_container_width=True, key="nav_logout", type="secondary"):
            logout()
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)


def main():
    from modules.auth import usuario_logado, eh_admin, supabase_configurado, restaurar_sessao
    import extra_streamlit_components as stx
    from datetime import datetime, timedelta

    if not supabase_configurado():
        st.error(
            "⚠️ **Supabase não configurado.**\n\n"
            "Adicione `SUPABASE_URL` e `SUPABASE_ANON_KEY` nos segredos do Streamlit "
            "(Settings → Secrets) ou no arquivo `.env`.",
            icon="🔒",
        )
        st.stop()

    # ── CookieManager sempre renderizado ─────────────────────────────────────
    # st.context.cookies não funciona no Community Cloud (CDN remove headers HTTP).
    # CookieManager usa JavaScript/WebSocket (document.cookie) — funciona sempre.
    # Só dispara reruns quando o valor do cookie muda — navegação normal é rápida.
    cm = stx.CookieManager(key="__le_cm")

    # ── Restaura sessão do cookie ─────────────────────────────────────────────
    if "user" not in st.session_state:
        if not st.session_state.get("_cm_init_done"):
            # Primeiro render: CookieManager ainda não inicializou (retorna None).
            # Seta flag e para — CookieManager dispara rerun automático (~100-300ms).
            st.session_state["_cm_init_done"] = True
            st.stop()
        rt = cm.get(_COOKIE_NAME)
        if rt:
            restaurar_sessao(rt)

    # ── Persiste sheets_creds no Supabase se veio de redirect OAuth ──────────
    if "user" in st.session_state and st.session_state.pop("_pending_sheets_save", False):
        if st.session_state.get("sheets_creds"):
            from modules.database import salvar_configuracoes
            salvar_configuracoes({"google_sheets_creds": st.session_state["sheets_creds"]})

    # ── Escreve cookie após login ─────────────────────────────────────────────
    if st.session_state.get("_pending_rt"):
        rt = st.session_state.pop("_pending_rt")
        cm.set(_COOKIE_NAME, rt, expires_at=datetime.now() + timedelta(days=30))

    # ── Apaga cookie no logout ────────────────────────────────────────────────
    if st.session_state.get("_do_logout_cookie"):
        try:
            cm.delete(_COOKIE_NAME)
        except Exception:
            pass
        st.session_state.pop("_do_logout_cookie", None)

    user = usuario_logado()

    if not user:
        pagina_login()
        return

    _sidebar()

    page = st.session_state.get("page", "busca")

    if page == "busca":
        pagina_busca()
    elif page == "historico":
        pagina_historico()
    elif page == "configuracoes":
        pagina_configuracoes()
    elif page == "admin":
        if eh_admin():
            pagina_admin()
        else:
            st.error("Acesso negado.")
    else:
        pagina_busca()


if __name__ == "__main__" or True:
    main()
