"""
Módulo de exportação de resultados para CSV e Excel.
"""

import csv
import os
from datetime import datetime
from pathlib import Path

try:
    from rich.console import Console
    from rich.table import Table
    _console = Console()
    _RICH = True
except Exception:
    _RICH = False
    _console = None

COLUNAS_PADRAO = [
    ("nome", "Nome"),
    ("telefone", "Telefone"),
    ("telefone2", "Telefone 2"),
    ("telefone_internacional", "Telefone Intl."),
    ("email", "E-mail"),
    ("endereco", "Endereço"),
    ("municipio", "Município"),
    ("uf", "UF"),
    ("cep", "CEP"),
    ("site", "Site"),
    ("maps_url", "Google Maps"),
    ("avaliacao", "Avaliação"),
    ("total_avaliacoes", "Nº Avaliações"),
    ("status_funcionamento", "Status"),
    ("cnpj", "CNPJ"),
    ("porte", "Porte"),
    ("data_abertura", "Data Abertura"),
    ("cidade_busca", "Cidade Buscada"),
    ("estado_busca", "Estado Buscado"),
    ("termo_busca", "Termo Extra"),
    ("fonte", "Fonte"),
]


def _gerar_nome_arquivo(prefixo: str, cidade: str, extensao: str) -> str:
    agora = datetime.now().strftime("%Y%m%d_%H%M%S")
    cidade_slug = cidade.lower().replace(" ", "_").replace("/", "-")
    return f"{prefixo}_{cidade_slug}_{agora}.{extensao}"


def exportar_csv(
    resultados: list[dict],
    cidade: str = "",
    caminho: str = None,
    diretorio: str = "resultados",
) -> str:
    """
    Exporta resultados para CSV.

    Retorna o caminho do arquivo gerado.
    """
    Path(diretorio).mkdir(parents=True, exist_ok=True)

    if caminho is None:
        caminho = os.path.join(diretorio, _gerar_nome_arquivo("prospecao", cidade, "csv"))

    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        # Cabeçalho com nomes amigáveis
        cabecalho = [label for _, label in COLUNAS_PADRAO]
        writer.writerow(cabecalho)

        for r in resultados:
            linha = [r.get(col, "") for col, _ in COLUNAS_PADRAO]
            writer.writerow(linha)

    print(f"[bold green]✓ CSV salvo:[/bold green] {caminho}")
    return caminho


def exportar_excel(
    resultados: list[dict],
    cidade: str = "",
    caminho: str = None,
    diretorio: str = "resultados",
) -> str:
    """
    Exporta resultados para Excel (.xlsx) com formatação.

    Retorna o caminho do arquivo gerado.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[red]openpyxl não instalado. Execute: pip install openpyxl[/red]")
        return exportar_csv(resultados, cidade, None, diretorio)

    Path(diretorio).mkdir(parents=True, exist_ok=True)

    if caminho is None:
        caminho = os.path.join(diretorio, _gerar_nome_arquivo("prospecao", cidade, "xlsx"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospecção"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")

    cabecalho = [label for _, label in COLUNAS_PADRAO]
    for col_idx, label in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20

    # Estilo alternado de linhas
    fill_par = PatternFill("solid", fgColor="D6E4F0")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, r in enumerate(resultados, start=2):
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        for col_idx, (col, _) in enumerate(COLUNAS_PADRAO, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(col, ""))
            cell.fill = fill
            if col in ("site", "maps_url") and r.get(col):
                cell.hyperlink = r[col]
                cell.font = Font(color="0563C1", underline="single")

    # Ajusta largura das colunas automaticamente
    for col_idx in range(1, len(cabecalho) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(resultados) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Congela cabeçalho
    ws.freeze_panes = "A2"

    wb.save(caminho)
    print(f"[bold green]✓ Excel salvo:[/bold green] {caminho}")
    return caminho


def mostrar_preview(resultados: list[dict], limite: int = 10) -> None:
    """Mostra uma prévia dos resultados no terminal."""
    if not resultados:
        print("Nenhum resultado para exibir.")
        return

    if _RICH:
        table = Table(title=f"Prévia — {len(resultados)} resultados", show_lines=True)
        table.add_column("Nome", style="bold", max_width=35)
        table.add_column("Telefone", style="cyan")
        table.add_column("Cidade", style="green")
        table.add_column("Site", style="dim", max_width=30)
        for r in resultados[:limite]:
            table.add_row(
                r.get("nome", "—"),
                r.get("telefone", "—") or r.get("telefone2", "—") or "—",
                r.get("municipio", "") or r.get("cidade_busca", "—"),
                r.get("site", "—") or "—",
            )
        if len(resultados) > limite:
            table.caption = f"... e mais {len(resultados) - limite} resultados"
        _console.print(table)
    else:
        print(f"\nPrévia — {len(resultados)} resultados:")
        for r in resultados[:limite]:
            print(f"  {r.get('nome','—')} | {r.get('telefone','—')} | {r.get('municipio','—')}")
